import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
import os

def clear_empty_cell_formatting(file_path, output_path):
    print(f"Processing Excel file: {file_path}")
    print(f"Original file size: {os.path.getsize(file_path) / (1024 * 1024):.2f} MB")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Initialize counters
        cleared_cells = 0
        total_cells = 0
        
        # Iterate through each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nProcessing worksheet: {sheet_name}")
            
            if ws.max_row and ws.max_column:
                sheet_rows = ws.max_row
                sheet_cols = ws.max_column
                print(f"Rows: {sheet_rows}, Columns: {sheet_cols}")
                
                total_cells += sheet_rows * sheet_cols
                
                # Iterate through all cells in the used range
                for row in ws.iter_rows(min_row=1, max_row=sheet_rows, min_col=1, max_col=sheet_cols):
                    for cell in row:
                        # Check if cell is empty (no value) and has formatting
                        if cell.value is None and (
                            cell.fill != PatternFill() or
                            cell.border != Border() or
                            cell.font != Font() or
                            cell.alignment != Alignment()
                        ):
                            # Clear formatting
                            cell.fill = PatternFill()
                            cell.border = Border()
                            cell.font = Font()
                            cell.alignment = Alignment()
                            cleared_cells += 1
                            
            else:
                print("Worksheet is empty or has no defined dimensions")
        
        # Save the modified workbook
        wb.save(output_path)
        print(f"\nSaved modified file to: {output_path}")
        print(f"New file size: {os.path.getsize(output_path) / (1024 * 1024):.2f} MB")
        print(f"Cleared formatting from {cleared_cells} empty cells out of {total_cells} total cells")
        
    except Exception as e:
        print(f"Error processing file: {e}")

if __name__ == "__main__":
    input_file = r"C:\Users\Greg\Desktop\file.xlsx"  # Update with correct file name, e.g., Inventory_2025.xlsx
    output_file = r"C:\Users\Greg\Desktop\file_cleaned.xlsx"  # Output file name
    if os.path.exists(input_file):
        clear_empty_cell_formatting(input_file, output_file)
    else:
        print(f"File not found: {input_file}")