import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
import os

def analyze_excel_file(file_path):
    print(f"Analyzing Excel file: {file_path}")
    print(f"File size: {os.path.getsize(file_path) / (1024 * 1024):.2f} MB")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, read_only=False)
        
        # 1. Count worksheets
        sheet_count = len(wb.sheetnames)
        print(f"Number of worksheets: {sheet_count}")
        
        # Initialize counters for analysis
        total_cells = 0
        formatted_cells = 0
        empty_formatted_cells = 0
        max_row = 0
        max_col = 0
        
        # Analyze each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nAnalyzing worksheet: {sheet_name}")
            
            # Get dimensions of used range
            if ws.max_row and ws.max_column:
                sheet_rows = ws.max_row
                sheet_cols = ws.max_column
                print(f"Rows: {sheet_rows}, Columns: {sheet_cols}")
                
                max_row = max(max_row, sheet_rows)
                max_col = max(max_col, sheet_cols)
                total_cells += sheet_rows * sheet_cols
                
                # Check for formatting and empty cells
                for row in ws.iter_rows(min_row=1, max_row=sheet_rows, min_col=1, max_col=sheet_cols):
                    for cell in row:
                        # Check if cell has formatting
                        has_formatting = (
                            cell.fill != PatternFill() or
                            cell.border != Border() or
                            cell.font != Font() or
                            cell.alignment != Alignment()
                        )
                        if has_formatting:
                            formatted_cells += 1
                            # Check if cell is empty but formatted
                            if cell.value is None:
                                empty_formatted_cells += 1
                            
            else:
                print("Worksheet is empty or has no defined dimensions")
        
        # Report findings
        print("\nSummary of Findings:")
        print(f"Total cells across all sheets: {total_cells}")
        print(f"Cells with formatting: {formatted_cells} ({(formatted_cells/total_cells)*100:.2f}% of total)")
        print(f"Empty cells with formatting: {empty_formatted_cells} ({(empty_formatted_cells/total_cells)*100:.2f}% of total)")
        print(f"Maximum rows used: {max_row}")
        print(f"Maximum columns used: {max_col}")
        
        # Check for embedded objects (approximation via comments or charts)
        chart_count = sum(len(ws._charts) for ws in wb.worksheets)
        comment_count = sum(sum(1 for cell in row if cell.comment) for row in ws.rows)
        print(f"Number of charts: {chart_count}")
        print(f"Number of comments: {comment_count}")
        
        # Check for conditional formatting
        cf_count = sum(len(ws.conditional_formatting) for ws in wb.worksheets)
        print(f"Number of conditional formatting rules: {cf_count}")
        
    except Exception as e:
        print(f"Error analyzing file: {e}")

if __name__ == "__main__":
    file_path = r"C:\Users\Greg\Desktop\file.xlsx"  # Update with the correct Excel file name
    if os.path.exists(file_path):
        analyze_excel_file(file_path)
    else:
        print(f"File not found: {file_path}")